from openpyxl import Workbook

wb = Workbook()
wb.active.title = "Sheet"
ws = wb["Sheet"]

ws['A1'] = 'data1'
ws['A2'] = 'data2'

wb.save('result.xlsx')